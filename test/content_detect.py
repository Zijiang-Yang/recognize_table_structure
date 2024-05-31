import openpyxl

# 加载xlsx文件
workbook = openpyxl.load_workbook('./assets/data/single/惠享e生 慢病百万医疗（三高版）费率 20220208.xlsx')

worksheet = workbook["Sheet2"]

merged_cells_ranges = worksheet.merged_cells.ranges

# 打印出每个合并单元格的范围
for merged_range in merged_cells_ranges:
    print(f"Merged cell range: {merged_range}")
    # 你也可以通过merged_range的start_row, start_column, end_row, end_column属性获取详细信息
    print(f"Start row: {merged_range.min_row}, Start column: {merged_range.min_col}, End row: {merged_range.max_row}, End column: {merged_range.max_col}")

for row in worksheet.iter_rows():
        for cell in row:
            # 输出cell的内容，行号和列号
            print(f'Row {cell.row}, Column {cell.column}, Value {cell.value}')